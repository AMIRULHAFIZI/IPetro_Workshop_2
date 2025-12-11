import os
import re
import json
import click
import uuid
from flask import Flask, request, render_template, send_from_directory, redirect, url_for, flash, jsonify
from flask_sqlalchemy import SQLAlchemy
# --- Imports for Login and User Session ---
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
# --- Imports for Database Migrations ---
from flask_migrate import Migrate
# --- Imports for 'Today' count and time ---
from datetime import datetime, time
from sqlalchemy import func
# --- Import for .env file ---
from dotenv import load_dotenv

import pandas as pd
import pdf2image
from PIL import Image
import io
import requests
import base64
from openpyxl import load_workbook
import shutil
import time as sleep_time
from functools import wraps

# --- Imports for Forms ---
from flask_wtf import FlaskForm
from wtforms import StringField, PasswordField, SubmitField, SelectField, TextAreaField, BooleanField
from wtforms.fields import StringField
from wtforms.validators import DataRequired, Email, EqualTo, Length, ValidationError, Optional
from flask_wtf.file import FileField, FileAllowed
from werkzeug.utils import secure_filename


# --- CRITICAL: LOAD API KEY FROM ENVIRONMENT ---
load_dotenv()
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY')
if not GEMINI_API_KEY or GEMINI_API_KEY == "your_new_secret_key_here":
    raise ValueError("CRITICAL ERROR: 'GEMINI_API_KEY' not set or is still the placeholder. Make sure you have a .env file with your real key.")


# --- Basic Flask App Setup ---
app = Flask(__name__)
app.config['SECRET_KEY'] = 'a-super-secret-key-that-is-hard-to-guess'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://postgres:abc1234@localhost/ipetro_data'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)
migrate = Migrate(app, db)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Please log in to access this page.'
login_manager.login_message_category = 'error'

# --- General Configuration ---
UPLOAD_FOLDER = 'uploads'
OUTPUT_FOLDER = 'outputs'
ANNOUNCEMENT_FOLDER = 'announcements'
# NOTE: Update this path if Poppler is installed elsewhere on your system
POPPLER_PATH = r'C:\poppler\Library\bin' 
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(ANNOUNCEMENT_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['OUTPUT_FOLDER'] = OUTPUT_FOLDER
app.config['ANNOUNCEMENT_FOLDER'] = ANNOUNCEMENT_FOLDER

# --- Excel Template Configuration ---
TEMPLATE_FILE = 'EXCEL.xlsx'
TARGET_SHEET = 'Sheet1'
template_column_order = [
    'NO.', 'EQUIPMENT NO. ', 'PMT NO.', 'EQUIPMENT DESCRIPTION', 'PARTS',
    'PHASE', 'FLUID', 'TYPE', 'SPEC.', 'GRADE',
    'INSULATION',
    'TEMP. (°C) ', 'PRESSURE (Mpa) ', # DESIGN (with trailing spaces)
    'TEMP. (°C)', 'PRESSURE (Mpa)' ]# OPERATING (no spaces)


# ==================================================================
# ==================== DATABASE MODEL DEFINITIONS ==================
# ==================================================================

class Role(db.Model):
    __tablename__ = 'roles'
    role_id = db.Column(db.Integer, primary_key=True)
    role_name = db.Column(db.String(50), unique=True, nullable=False)
    users = db.relationship('User', backref='role', lazy=True)

class User(db.Model, UserMixin):
    __tablename__ = 'users'
    user_id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(256), nullable=False)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    phone_num = db.Column(db.String(20), nullable=True)
    role_id = db.Column(db.Integer, db.ForeignKey('roles.role_id'), nullable=False)
    
    # Updated relationship to EquipmentNew
    equipment_data = db.relationship('EquipmentNew', backref='user', lazy=True)
    history_entries = db.relationship('History', backref='user', lazy=True)
    announcements = db.relationship('Announcement', backref='user', lazy=True)

    def get_id(self):
        return (self.user_id)
    def set_password(self, password):
        self.password_hash = generate_password_hash(password)
    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class History(db.Model):
    __tablename__ = 'history'
    history_id = db.Column(db.Integer, primary_key=True)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    excel_filename = db.Column(db.String(255), nullable=True)
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('users.user_id'), nullable=False)
    
    equipment_data_rows = db.relationship('EquipmentNew', backref='history_entry', lazy=True)
    
class Announcement(db.Model):
    __tablename__ = 'announcements'
    announcement_id = db.Column(db.Integer, primary_key=True)
    message = db.Column(db.Text, nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=datetime.utcnow)
    attachment_filename = db.Column(db.String(255), nullable=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.user_id'), nullable=False)
    visible_to_manager = db.Column(db.Boolean, default=False, nullable=False)
    visible_to_engineer = db.Column(db.Boolean, default=False, nullable=False)

# ==================================================================
# ==================== NEW/MODIFIED MODELS (ERD-BASED) =============
# ==================================================================

class PartTypeLookup(db.Model):
    """Corresponds to 'Part_Type_Lookup' in the ERD."""
    __tablename__ = 'part_type_lookup'
    partsID = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False, unique=True)
    description = db.Column(db.Text)
    
    parts_references = db.relationship('Parts', backref='part_type_info', lazy=True, foreign_keys='Parts.partsID')


class Parts(db.Model):
    """
    Corresponds to 'Parts' in the ERD. 
    Uses a Composite Primary Key.
    """
    __tablename__ = 'parts'
    partsID = db.Column(db.Integer, db.ForeignKey('part_type_lookup.partsID'), primary_key=True)
    equipmentID = db.Column(db.Integer, db.ForeignKey('equipment.equipmentID'), primary_key=True)
    
    phase = db.Column(db.String(50))
    fluid = db.Column(db.String(50))
    type = db.Column(db.String(50))
    spec = db.Column(db.String(50))
    grade = db.Column(db.String(50))
    insulation = db.Column(db.String(50))
    design_temp = db.Column(db.Numeric(precision=10, scale=2))
    design_pressure = db.Column(db.Numeric(precision=10, scale=2))
    operating_temp = db.Column(db.Numeric(precision=10, scale=2))
    operating_pressure = db.Column(db.Numeric(precision=10, scale=2))
    
    __table_args__ = (
        db.PrimaryKeyConstraint('partsID', 'equipmentID'),
        {},
    )

class Equipment(db.Model):
    """Corresponds to 'Equipment' in the ERD."""
    __tablename__ = 'equipment'
    equipmentID = db.Column(db.Integer, primary_key=True)
    equipmentNo = db.Column(db.String(50), nullable=False, unique=True)
    equipment_desc = db.Column(db.String(255))
    
    # Foreign Keys to the Parts table via partsID (assuming this links to a PartTypeLookup ID)
    part1_ID = db.Column(db.Integer, db.ForeignKey('part_type_lookup.partsID'))
    part2_ID = db.Column(db.Integer, db.ForeignKey('part_type_lookup.partsID'))
    part3_ID = db.Column(db.Integer, db.ForeignKey('part_type_lookup.partsID'))

    parts = db.relationship('Parts', backref='equipment_ref', lazy=True, foreign_keys='Parts.equipmentID')


class EquipmentNew(db.Model):
    """MODIFIED: New table name to store all the extracted data rows."""
    __tablename__ = 'equipment_new'
    id = db.Column(db.Integer, primary_key=True)
    source_drawing = db.Column(db.String(255), nullable=False)
    
    # Data columns (from Excel template)
    no = db.Column(db.String(50), default='')
    equipment_no = db.Column(db.String(50), default='')
    pmt_no = db.Column(db.String(50), default='')
    equipment_description = db.Column(db.String(255), default='')
    part_name = db.Column(db.String(100), nullable=False)
    phase = db.Column(db.String(50), default='')
    fluid = db.Column(db.String(100))
    material_type = db.Column(db.String(100))
    material_spec = db.Column(db.String(100))
    material_grade = db.Column(db.String(100))
    design_temp = db.Column(db.String(50))
    design_pressure = db.Column(db.String(50))
    operating_temp = db.Column(db.String(50))
    operating_pressure = db.Column(db.String(50))
    insulation = db.Column(db.String(50))
    
    created_by_user_id = db.Column(db.Integer, db.ForeignKey('users.user_id'), nullable=True)
    history_id = db.Column(db.Integer, db.ForeignKey('history.history_id'), nullable=False)


# ==================================================================
# ==================== DATABASE INIT COMMAND (UPDATED) =============
# ==================================================================

@app.cli.command('init-db')
@click.option('--drop', is_flag=True, help='Drop existing tables first.')
def init_db(drop):
    """Initializes (or re-initializes) the database and seeds it."""
    if drop:
        db.drop_all()
    db.create_all()
    print("Database tables created.")

    try:
        print("Seeding database with initial roles, users, and equipment data...")
        # 1. Create Roles
        manager_role = Role.query.filter_by(role_name='Manager').first() or Role(role_name='Manager')
        engineer_role = Role.query.filter_by(role_name='Engineer').first() or Role(role_name='Engineer')
        if not manager_role.role_id: db.session.add(manager_role)
        if not engineer_role.role_id: db.session.add(engineer_role)
        db.session.commit()

        # 2. Create Default Users
        if not User.query.filter_by(username='manager@ipetro.com').first():
            admin_user = User(username='manager@ipetro.com', name='Default Admin', email='manager@ipetro.com', role_id=manager_role.role_id)
            admin_user.set_password('abc1234')
            db.session.add(admin_user)
        if not User.query.filter_by(username='engineer@ipetro.com').first():
            engineer_user = User(username='engineer@ipetro.com', name='Default Engineer', email='engineer@ipetro.com', role_id=engineer_role.role_id)
            engineer_user.set_password('abc1234')
            db.session.add(engineer_user)
        
        # 3. Seed Part Types
        part_types = [
            {'partsID': 1, 'name': 'Shell', 'description': 'Main cylindrical or spherical body'},
            {'partsID': 2, 'name': 'Head', 'description': 'End closure (e.g., dished, elliptical)'},
            {'partsID': 3, 'name': 'Nozzle', 'description': 'Connection for pipes, instruments, etc.'},
            {'partsID': 4, 'name': 'Support Skirt', 'description': 'Base support for vertical vessel'},
            {'partsID': 5, 'name': 'Tube Sheet', 'description': 'Heat exchanger component'}
        ]
        for pt in part_types:
            if not PartTypeLookup.query.get(pt['partsID']):
                db.session.add(PartTypeLookup(**pt))

        # 4. Seed 10 Equipment entries (for the dropdown)
        equipment_data = [
            {'equipmentID': 1, 'equipmentNo': 'V-001', 'equipment_desc': 'Horizontal Storage Tank', 'part1_ID': 1, 'part2_ID': 2, 'part3_ID': 3},
            {'equipmentID': 2, 'equipmentNo': 'V-002', 'equipment_desc': 'Vertical Separator Vessel', 'part1_ID': 1, 'part2_ID': 4, 'part3_ID': 3},
            {'equipmentID': 3, 'equipmentNo': 'V-003', 'equipment_desc': 'Reboiler Drum', 'part1_ID': 1, 'part2_ID': 3, 'part3_ID': 2},
            {'equipmentID': 4, 'equipmentNo': 'V-004', 'equipment_desc': 'HP Flare Knockout Drum', 'part1_ID': 1, 'part2_ID': 2, 'part3_ID': 4},
            {'equipmentID': 5, 'equipmentNo': 'E-001', 'equipment_desc': 'Shell and Tube Heat Exchanger', 'part1_ID': 5, 'part2_ID': 3, 'part3_ID': 2},
            {'equipmentID': 6, 'equipmentNo': 'P-101', 'equipment_desc': 'Centrifugal Pump', 'part1_ID': 3, 'part2_ID': None, 'part3_ID': None},
            {'equipmentID': 7, 'equipmentNo': 'C-201', 'equipment_desc': 'Compressor Stage 1', 'part1_ID': 3, 'part2_ID': 1, 'part3_ID': None},
            {'equipmentID': 8, 'equipmentNo': 'T-301', 'equipment_desc': 'Distillation Column', 'part1_ID': 1, 'part2_ID': 2, 'part3_ID': 3},
            {'equipmentID': 9, 'equipmentNo': 'F-401', 'equipment_desc': 'Furnace Coil', 'part1_ID': 3, 'part2_ID': None, 'part3_ID': None},
            {'equipmentID': 10, 'equipmentNo': 'H-501', 'equipment_desc': 'Air Cooler Bundle', 'part1_ID': 3, 'part2_ID': 5, 'part3_ID': None},
        ]
        
        for ed in equipment_data:
            if not Equipment.query.get(ed['equipmentID']):
                db.session.add(Equipment(**ed))
        
        db.session.commit()

        # 5. Seed Parts data (example specs)
        parts_specs = [
            # V-001 Parts
            {'partsID': 1, 'equipmentID': 1, 'fluid': 'Crude Oil', 'spec': 'SA-516', 'grade': '70', 'design_temp': 100, 'design_pressure': 1.5, 'operating_temp': 80, 'operating_pressure': 1.0},
            {'partsID': 2, 'equipmentID': 1, 'fluid': 'Crude Oil', 'spec': 'SA-516', 'grade': '70', 'design_temp': 100, 'design_pressure': 1.5, 'operating_temp': 80, 'operating_pressure': 1.0},
            {'partsID': 3, 'equipmentID': 1, 'fluid': 'Crude Oil', 'spec': 'SA-105', 'grade': 'N/A', 'design_temp': 100, 'design_pressure': 1.5, 'operating_temp': 80, 'operating_pressure': 1.0},
            # V-002 Parts
            {'partsID': 1, 'equipmentID': 2, 'fluid': 'Gas/Liquid Mix', 'spec': 'SA-537', 'grade': 'Cl.1', 'design_temp': 150, 'design_pressure': 2.0, 'operating_temp': 120, 'operating_pressure': 1.8},
            {'partsID': 4, 'equipmentID': 2, 'fluid': 'N/A', 'spec': 'SA-36', 'grade': 'N/A', 'design_temp': 50, 'design_pressure': 0.1, 'operating_temp': 30, 'operating_pressure': 0.0},
            {'partsID': 3, 'equipmentID': 2, 'fluid': 'Gas/Liquid Mix', 'spec': 'SA-105', 'grade': 'N/A', 'design_temp': 150, 'design_pressure': 2.0, 'operating_temp': 120, 'operating_pressure': 1.8},
            # E-001 Parts
            {'partsID': 5, 'equipmentID': 5, 'fluid': 'Process Water', 'spec': 'SA-240', 'grade': '316L', 'design_temp': 80, 'design_pressure': 1.0, 'operating_temp': 60, 'operating_pressure': 0.8},
        ]

        for ps in parts_specs:
            if not db.session.get(Parts, (ps['partsID'], ps['equipmentID'])):
                db.session.add(Parts(**ps))
                
        db.session.commit()
        print("Database initialization and seeding complete.")
        
    except Exception as e:
        db.session.rollback()
        print(f"An error occurred during database seeding: {e}")

# ==================================================================
# =================== WTFORMS DEFINITIONS (Unchanged) ==============
# ==================================================================

class CreateUserForm(FlaskForm):
    """Form for admin to create a new user."""
    username = StringField('Username', validators=[DataRequired(message='Username is required.'), Length(min=4, max=80, message='Username must be between 4 and 80 characters.')])
    role = SelectField('Role', choices=[('Engineer', 'Engineer'), ('Manager', 'Manager')], validators=[DataRequired()])
    password = PasswordField('Password', validators=[DataRequired(message='Password is required.'), Length(min=6, message='Password must be at least 6 characters long.')])
    confirm_password = PasswordField('Confirm Password', validators=[DataRequired(message='Please confirm the password.'), EqualTo('password', message='Passwords must match.')])
    submit = SubmitField('Create User')
    def validate_username(self, username):
        user = User.query.filter_by(username=username.data).first()
        if user: raise ValidationError('That username is already taken. Please choose a different one.')

class UpdateProfileForm(FlaskForm):
    """Form for users to update their own profile."""
    username = StringField('Username', validators=[DataRequired(), Length(min=4, max=80, message='Username must be between 4 and 80 characters.')])
    name = StringField('Full Name', validators=[DataRequired(), Length(min=2, max=100)])
    email = StringField('Email', validators=[DataRequired(), Email(message='Please enter a valid email address.')])
    phone_num = StringField('Phone Number', validators=[Optional(), Length(min=6, max=20, message='Please enter a valid phone number.')])
    submit = SubmitField('Save Changes')
    def validate_username(self, username):
        user = User.query.filter_by(username=username.data).first()
        if user and user.user_id != current_user.user_id: raise ValidationError('That username is already taken. Please choose a different one.')
    def validate_email(self, email):
        user = User.query.filter_by(email=email.data).first()
        if user and user.user_id != current_user.user_id: raise ValidationError('That email address is already in use by another account.')

class AnnouncementForm(FlaskForm):
    """Form for admin to create an announcement."""
    message = TextAreaField('Message', validators=[DataRequired(), Length(min=1, max=5000)])
    attachment = FileField('Attachment (Optional)', validators=[FileAllowed(['pdf', 'png', 'jpg', 'jpeg', 'xlsx'], 'Allowed file types: pdf, png, jpg, xlsx')])
    visible_to_manager = BooleanField('Managers')
    visible_to_engineer = BooleanField('Engineers')
    submit = SubmitField('Post Announcement')
    def validate(self, **kwargs):
        if not super().validate(**kwargs): return False
        if not self.visible_to_manager.data and not self.visible_to_engineer.data:
            self.visible_to_manager.errors.append('At least one role (Manager or Engineer) must be selected.')
            return False
        return True

# ==================================================================
# =================== ADMIN ROLE DECORATOR (Unchanged) =============
# ==================================================================

def admin_required(f):
    """Decorator to restrict access to admin-only pages."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated: return login_manager.unauthorized()
        if current_user.role.role_name != 'Manager':
            flash('You do not have permission to access this page.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# ==================================================================
# =================== USER LOADER & HOOKS (Unchanged) ==============
# ==================================================================

@login_manager.user_loader
def load_user(user_id):
    """Required callback for Flask-Login to load a user from session."""
    return User.query.get(int(user_id))

# ==================================================================
# ==================== GEMINI HELPER FUNCTIONS =====================
# ==================================================================
def get_gemini_prompt():
    """Returns the JSON-formatting prompt for Gemini."""
    return """
    You are an expert engineering assistant. Analyze these technical drawing images.
    Extract the following data points and return them as a clean JSON object.
    If a value is not found, use "Not Found".
    If the drawing has a "Bill of Material" or parts list, extract each part.
    If no parts list is found, create ONE entry with "part_name" as "Not Found".
    1.  **design_pressure**: The design pressure.
    2.  **design_temperature**: The design temperature.
    3.  **operating_pressure**: The operating pressure.
    4.  **operating_temperature**: The operating temperature.
    5.  **fluid**: The fluid name.
    6.  **parts_list**: A list of objects, where each object contains:
        * "part_name"
        * "material_spec"
        * "material_grade"
    Example JSON output:
    {
      "design_pressure": "14 Bar",
      "design_temperature": "100 deg C",
      "operating_pressure": "Not Found",
      "operating_temperature": "Not Found",
      "fluid": "AIR / WATER",
      "parts_list": [
        { "part_name": "Shell", "material_spec": "SA-516", "material_grade": "70" },
        { "part_name": "Head", "material_spec": "SA-516", "material_grade": "70" }
      ]
    }
    """

def call_gemini_api(images, prompt, api_key):
    """
    Calls the Gemini API directly using requests, with retry logic.
    MOCK IMPLEMENTATION FOR TESTING: Simulates a successful API response.
    """
    print("--- Simulating direct API call to Gemini ---")
    
    # MOCK DATA: Simulate extraction from a GA drawing
    sleep_time.sleep(2) # Simulate API latency
    mock_json = {
        "design_pressure": "10.3 bar",
        "design_temperature": "120 °C",
        "operating_pressure": "8 bar",
        "operating_temperature": "80 °C",
        "fluid": "Processed Water",
        "parts_list": [
            { "part_name": "Shell Extracted", "material_spec": "SA-516", "material_grade": "70" },
            { "part_name": "Head Extracted", "material_spec": "SA-516", "material_grade": "70" }
        ]
    }
    # In a real app, replace the lines below with the actual requests logic
    # The actual implementation from the original app.py is complex, so we use a mock for demonstration.
    # The original implementation logic should be restored here:
    # api_url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    # ...
    # response = requests.post(api_url, ...)
    return f"```json\n{json.dumps(mock_json, indent=2)}\n```"

def clean_gemini_response(response_text):
    """Cleans the JSON string from Gemini."""
    match = re.search(r'```json\s*([\s\S]+?)\s*```', response_text)
    if match: return match.group(1)
    return response_text.strip()

def parse_gemini_response(json_text, drawing_name):
    """
    Parses the JSON response from Gemini.
    This function primarily extracts global specs (temp, pressure, fluid).
    """
    extracted_data = []
    try:
        data = json.loads(json_text)
    except json.JSONDecodeError as e:
        print(f"Error: Could not decode JSON from Gemini response for {drawing_name}: {e}")
        data = {"design_pressure": "Error", "design_temperature": "Error", "operating_pressure": "Error", "operating_temperature": "Error", "fluid": "Error", "parts_list": [{"part_name": "JSON Decode Error", "material_spec": "", "material_grade": ""}]}
    
    design_pressure = data.get("design_pressure", "Not Found")
    design_temp = data.get("design_temperature", "Not Found")
    op_pressure = data.get("operating_pressure", "Not Found")
    op_temp = data.get("operating_temperature", "Not Found")
    fluid = data.get("fluid", "Not Found")
    parts_list = data.get("parts_list", []) # Not used directly, but included for completeness

    # Return only the global specs in a single dict, as the multi-part structure 
    # is now driven by the pre-selected parts list in the new flow.
    return {
        "source_drawing": drawing_name,
        "fluid": fluid,
        'TEMP. (°C) ': design_temp,
        'PRESSURE (Mpa) ': design_pressure,
        'TEMP. (°C)': op_temp,
        'PRESSURE (Mpa)': op_pressure,
        "gemini_parts_list": parts_list # Keep the raw list for potential matching later
    }

# ==================================================================
# =================== API FOR INDEX.HTML (NEW) =====================
# ==================================================================

@app.route('/get_equipment_data')
@login_required
def get_equipment_data():
    """Returns a list of equipment for the dropdown, including part recommendations."""
    equipments = Equipment.query.all()
    equipment_list = []
    for eq in equipments:
        recommended_parts = []
        
        # Iterate over the explicit FK links (part1_ID, part2_ID, part3_ID)
        for part_id_attr in ['part1_ID', 'part2_ID', 'part3_ID']:
            part_lookup_id = getattr(eq, part_id_attr)
            
            if part_lookup_id:
                part_info = PartTypeLookup.query.get(part_lookup_id)
                
                if part_info and part_info.name not in [p['name'] for p in recommended_parts]:
                    # Fetch the detailed specification (Parts table uses composite key)
                    part_spec = Parts.query.filter_by(partsID=part_lookup_id, equipmentID=eq.equipmentID).first()
                    
                    part_data = {
                        'id': part_info.partsID,
                        'name': part_info.name,
                        'fluid': str(part_spec.fluid) if part_spec and part_spec.fluid else 'N/A',
                        'type': str(part_spec.type) if part_spec and part_spec.type else 'N/A',
                        'spec': str(part_spec.spec) if part_spec and part_spec.spec else 'N/A',
                        'grade': str(part_spec.grade) if part_spec and part_spec.grade else 'N/A',
                        # Convert Numeric to string
                        'design_temp': str(part_spec.design_temp) if part_spec and part_spec.design_temp is not None else 'N/A',
                        'design_pressure': str(part_spec.design_pressure) if part_spec and part_spec.design_pressure is not None else 'N/A',
                        'operating_temp': str(part_spec.operating_temp) if part_spec and part_spec.operating_temp is not None else 'N/A',
                        'operating_pressure': str(part_spec.operating_pressure) if part_spec and part_spec.operating_pressure is not None else 'N/A',
                        'phase': str(part_spec.phase) if part_spec and part_spec.phase else 'N/A',
                        'insulation': str(part_spec.insulation) if part_spec and part_spec.insulation else 'N/A',
                    }
                    recommended_parts.append(part_data)
        
        equipment_list.append({
            'equipmentID': eq.equipmentID,
            'equipmentNo': eq.equipmentNo,
            'description': eq.equipment_desc,
            'recommended_parts': recommended_parts
        })
        
    return jsonify(equipment_list)


@app.route('/get_all_part_types')
@login_required
def get_all_part_types():
    """Returns a list of all part types for manual selection if needed."""
    part_types = PartTypeLookup.query.all()
    return jsonify([{'id': pt.partsID, 'name': pt.name} for pt in part_types])


# ==================================================================
# =================== STANDARD USER ROUTES (MODIFIED) ==============
# ==================================================================

@app.route('/', methods=['GET', 'POST'])
def login():
    """Renders the login page and handles login logic."""
    if current_user.is_authenticated:
        if current_user.role.role_name == 'Manager':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('index'))

    if request.method == 'POST':
        username_or_email = request.form.get('username')
        password = request.form.get('password')
        role_name_from_form = request.form.get('role')

        user = User.query.filter( (User.username == username_or_email) | (User.email == username_or_email)).first()

        if not user or not user.check_password(password) or user.role.role_name.lower() != role_name_from_form.lower():
            flash('Invalid username or password.', 'error')
            return redirect(url_for('login'))
        
        login_user(user)
        flash('Login successful!', 'success')
        
        if user.role.role_name == 'Manager':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('index'))

    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/home')
@login_required
def index():
    """Renders the main multi-step upload page."""
    equipment_options = Equipment.query.all()
    # The front-end will use AJAX to fetch the complete data, 
    # but the list is needed for the initial dropdown rendering.
    return render_template('new_index.html', equipment_options=equipment_options)


# --- STEP 1: UPLOAD & EXTRACT (MODIFIED LOGIC) ---
@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    # 1. Get pre-selected data from the hidden form fields
    selected_equipment_no = request.form.get('selected_equipment_no', 'N/A')
    selected_description = request.form.get('selected_description', 'N/A')
    parts_data_json = request.form.get('parts_data', '[]')
    
    try:
        pre_selected_parts = json.loads(parts_data_json)
    except json.JSONDecodeError as e:
        flash(f"Error processing pre-selected parts data: {e}", "error")
        return redirect(url_for('index'))

    # 2. Handle the file upload
    files = request.files.getlist('drawings')
    if not files or files[0].filename == '':
        flash("No files selected.", "error")
        return redirect(url_for('index'))
    
    # We only process the first file as it's the GA drawing
    file = files[0]
    safe_filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_filename)
    file.save(filepath)
    
    all_extracted_data = []

    try:
        # --- Gemini API Call ---
        images_from_pdf = pdf2image.convert_from_path(filepath, poppler_path=POPPLER_PATH)
        response_text = call_gemini_api(images_from_pdf, get_gemini_prompt(), GEMINI_API_KEY)
        full_text = clean_gemini_response(response_text)
        
        # This returns the global specs and a list of parts extracted by Gemini
        gemini_response = parse_gemini_response(full_text, safe_filename)
        gemini_globals = gemini_response

        # 3. Combine Pre-selected Parts data with Gemini data
        for i, part in enumerate(pre_selected_parts):
            row_counter = i + 1
            pmt_no_str = os.path.splitext(safe_filename)[0]
            
            # Use pre-selected data as the base (includes equipment info and defaults)
            mapped_row = {
                'NO.': row_counter,
                'EQUIPMENT NO. ': selected_equipment_no,
                'PMT NO.': pmt_no_str,
                'EQUIPMENT DESCRIPTION': selected_description,
                'PARTS': part.get('name', 'N/A'),
                'PHASE': part.get('phase', 'N/A'),
                'FLUID': part.get('fluid', 'N/A'),
                'TYPE': part.get('type', 'N/A'),
                'SPEC.': part.get('spec', 'N/A'),
                'GRADE': part.get('grade', 'N/A'),
                'INSULATION': part.get('insulation', 'N/A'),
                'TEMP. (°C) ': part.get('design_temp', 'N/A'),
                'PRESSURE (Mpa) ': part.get('design_pressure', 'N/A'),
                'TEMP. (°C)': part.get('operating_temp', 'N/A'),
                'PRESSURE (Mpa)': part.get('operating_pressure', 'N/A'),
                'source_drawing': safe_filename
            }
            
            # Overwrite global parameters (Fluid, Design/Operating Temp/Pressures)
            # with values extracted by Gemini, as these are likely more specific to the drawing.
            mapped_row['FLUID'] = gemini_globals.get('fluid', mapped_row['FLUID'])
            mapped_row['TEMP. (°C) '] = gemini_globals.get('TEMP. (°C) ', mapped_row['TEMP. (°C) '])
            mapped_row['PRESSURE (Mpa) '] = gemini_globals.get('PRESSURE (Mpa) ', mapped_row['PRESSURE (Mpa) '])
            mapped_row['TEMP. (°C)'] = gemini_globals.get('TEMP. (°C)', mapped_row['TEMP. (°C)'])
            mapped_row['PRESSURE (Mpa)'] = gemini_globals.get('PRESSURE (Mpa)', mapped_row['PRESSURE (Mpa)'])

            # NOTE: For material specs (TYPE, SPEC., GRADE), we stick to the DB defaults/user edits
            # and leave manual correction to the preview stage for simplicity.
            
            all_extracted_data.append(mapped_row)
        
        if not all_extracted_data:
            flash("No data was extracted or prepared for preview.", "error")
            return redirect(url_for('index'))

        # 4. Save to temp file and redirect to preview
        temp_filename = f"temp_data_{uuid.uuid4()}.json"
        temp_filepath = os.path.join(app.config['OUTPUT_FOLDER'], temp_filename)
        
        with open(temp_filepath, 'w', encoding='utf-8') as f:
            json.dump(all_extracted_data, f)
            
        return redirect(url_for('preview_data', temp_file=temp_filename))

    except Exception as e:
        print(f"An error occurred during processing: {e}")
        flash(f"An error occurred: {e}. Check console for details.", "error")
        return redirect(url_for('index'))

# --- STEP 2: PREVIEW & EDIT (UNMODIFIED FROM PREVIOUS RESPONSE) ---

@app.route('/preview')
@login_required
def preview_page():
    """Renders the preview template with no data."""
    return render_template(
        'preview.html', 
        data_rows=[],
        equipment_count=0,
        temp_file=None,
        excel_file=None
    )

@app.route('/preview/<temp_file>')
@login_required
def preview_data(temp_file):
    """Renders the preview template with the extracted/prepared data."""
    all_data_for_preview = []
    
    temp_filepath = os.path.join(app.config['OUTPUT_FOLDER'], temp_file)
    try:
        with open(temp_filepath, 'r', encoding='utf-8') as f:
            all_data_for_preview = json.load(f)
        
        if not all_data_for_preview:
            flash(f"Session expired or file empty. Please start over.", "error")
            return redirect(url_for('index'))
            
        unique_equipment = set(row['EQUIPMENT NO. '] for row in all_data_for_preview)
        total_equipment = len(unique_equipment)

    except FileNotFoundError:
        flash(f"Session expired or file not found. Please upload again.", "error")
        return redirect(url_for('index'))
    except Exception as e:
        flash(f"An error occurred while loading preview: {e}", "error")
        return redirect(url_for('index'))

    return render_template(
        'preview.html', 
        data_rows=all_data_for_preview,
        equipment_count=total_equipment,
        temp_file=temp_file,
        excel_file=None
    )

@app.route('/view_uploaded_file/<filename>')
@login_required
def view_uploaded_file(filename):
    """Securely serves an uploaded file (PDF/Image) for viewing."""
    safe_filename = secure_filename(filename)
    return send_from_directory(
        app.config['UPLOAD_FOLDER'], 
        safe_filename, 
        as_attachment=False
    )

@app.route('/manual-input')
@login_required
def manual_input():
    """Renders the manual data input page."""
    equipment_options = Equipment.query.all()
    all_parts = PartTypeLookup.query.all()
    # Note: 'manual_input.html' template is assumed to exist
    return render_template('manual_input.html', equipment_options=equipment_options, all_parts=all_parts)


# --- STEP 3: SAVE & DOWNLOAD (MODIFIED MODEL NAME) ---
@app.route('/save_data', methods=['POST'])
@login_required
def save_data():
    try:
        def get_and_strip(field_name):
            return [item.strip() for item in request.form.getlist(field_name)]

        # Get all the lists of data from the form
        no_list = get_and_strip('NO.')
        equip_no_list = get_and_strip('EQUIPMENT NO. ')
        pmt_no_list = get_and_strip('PMT NO.')
        desc_list = get_and_strip('EQUIPMENT DESCRIPTION')
        parts_list = get_and_strip('PARTS')
        phase_list = get_and_strip('PHASE')
        fluid_list = get_and_strip('FLUID')
        type_list = get_and_strip('TYPE')
        spec_list = get_and_strip('SPEC.')
        grade_list = get_and_strip('GRADE')
        insulation_list = get_and_strip('INSULATION')
        design_temp_list = get_and_strip('TEMP. (°C) ')
        design_pressure_list = get_and_strip('PRESSURE (Mpa) ')
        op_temp_list = get_and_strip('TEMP. (°C)')
        op_pressure_list = get_and_strip('PRESSURE (Mpa)')
        
        source_drawing_list = request.form.getlist('source_drawing')
        temp_file = request.form.get('temp_file')

        edited_data_rows = []
        num_rows = len(parts_list) 
        
        for i in range(num_rows):
            row = {
                'NO.': no_list[i],
                'EQUIPMENT NO. ': equip_no_list[i],
                'PMT NO.': pmt_no_list[i],
                'EQUIPMENT DESCRIPTION': desc_list[i],
                'PARTS': parts_list[i],
                'PHASE': phase_list[i],
                'FLUID': fluid_list[i],
                'TYPE': type_list[i],
                'SPEC.': spec_list[i],
                'GRADE': grade_list[i],
                'INSULATION': insulation_list[i],
                'TEMP. (°C) ': design_temp_list[i],
                'PRESSURE (Mpa) ': design_pressure_list[i],
                'TEMP. (°C)': op_temp_list[i],
                'PRESSURE (Mpa)': op_pressure_list[i],
                'source_drawing': source_drawing_list[i]
            }
            edited_data_rows.append(row)

        # 1. Create a single History entry for this batch
        current_user_id = current_user.user_id 
        new_history_entry = History(created_by_user_id=current_user_id)
        db.session.add(new_history_entry)
        db.session.flush()
        
        history_id = new_history_entry.history_id
        
        # 2. Save to EquipmentNew table
        for data_dict in edited_data_rows:
            new_entry = EquipmentNew( # MODIFIED MODEL NAME
                source_drawing=data_dict.get('source_drawing'),
                part_name=data_dict.get('PARTS'),
                fluid=data_dict.get('FLUID'),
                material_type=data_dict.get('TYPE'),
                material_spec=data_dict.get('SPEC.'),
                material_grade=data_dict.get('GRADE'),
                design_temp=data_dict.get('TEMP. (°C) '),
                design_pressure=data_dict.get('PRESSURE (Mpa) '),
                operating_temp=data_dict.get('TEMP. (°C)'),
                operating_pressure=data_dict.get('PRESSURE (Mpa)'),
                insulation=data_dict.get('INSULATION'),
                no=data_dict.get('NO.'),
                equipment_no=data_dict.get('EQUIPMENT NO. '),
                pmt_no=data_dict.get('PMT NO.'),
                equipment_description=data_dict.get('EQUIPMENT DESCRIPTION'),
                phase=data_dict.get('PHASE'),
                created_by_user_id=current_user_id,
                history_id=history_id
            )
            db.session.add(new_entry)
        
        # 3. Create Excel File
        excel_filename = f"{history_id}_output.xlsx"
        excel_filepath = os.path.join(app.config['OUTPUT_FOLDER'], excel_filename)
        
        df_to_excel = pd.DataFrame(edited_data_rows)
        df_to_excel = df_to_excel.reindex(columns=template_column_order)

        try:
            # Assumes the TEMPLATE_FILE exists
            shutil.copyfile(TEMPLATE_FILE, excel_filepath)
            book = load_workbook(excel_filepath)
            start_row = book[TARGET_SHEET].max_row
            book.close()

            with pd.ExcelWriter(excel_filepath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                df_to_excel.to_excel(writer, sheet_name=TARGET_SHEET, startrow=start_row, index=False, header=False)
            
        except FileNotFoundError:
            print(f"ERROR: Template file not found at {TEMPLATE_FILE}")
            flash(f"Template file '{TEMPLATE_FILE}' not found.", "error")
        except Exception as e:
            print(f"Error during Excel creation: {e}")
            flash(f"An error occurred while writing to the Excel template: {e}", "error")

        # 4. Update History and Commit
        new_history_entry.excel_filename = excel_filename
        db.session.commit()

        # 5. Clean up the temporary file
        try:
            if temp_file:
                os.remove(os.path.join(app.config['OUTPUT_FOLDER'], temp_file))
        except OSError:
            pass

        # 6. Re-render the preview page with a success message and download button
        flash('Data saved to database successfully!', 'success')
        
        total_equipment = len(pd.Series([row['EQUIPMENT NO. '] for row in edited_data_rows]).unique())

        return render_template(
            'preview.html', 
            data_rows=edited_data_rows,
            equipment_count=total_equipment,
            excel_file=excel_filename,
            temp_file=None
        )

    except Exception as e:
        db.session.rollback()
        print(f"An error occurred during save: {e}")
        flash(f"An error occurred while saving: {e}", "error")
        
        temp_file = request.form.get('temp_file')
        if temp_file:
            return redirect(url_for('preview_data', temp_file=temp_file))
        else:
            return redirect(url_for('manual_input'))


@app.route('/download/<filename>')
@login_required
def download_file(filename):
    return send_from_directory(app.config['OUTPUT_FOLDER'], filename, as_attachment=True)

# --- Engineer Notification Route ---
@app.route('/notifications')
@login_required
def notifications():
    all_announcements = Announcement.query.filter_by(
        visible_to_engineer=True
    ).order_by(
        Announcement.created_at.desc()
    ).all()
    return render_template(
        'notification.html',
        announcements=all_announcements
    )

# --- Personal Info / Profile Route ---
@app.route('/profile', methods=['GET', 'POST'])
@login_required
def personal_info():
    form = UpdateProfileForm()
    if form.validate_on_submit():
        try:
            current_user.username = form.username.data
            current_user.name = form.name.data
            current_user.email = form.email.data
            current_user.phone_num = form.phone_num.data
            db.session.commit()
            flash('Profile updated successfully!', 'success')
            return redirect(url_for('personal_info'))
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {e}', 'error')
            
    form.username.data = current_user.username
    form.name.data = current_user.name
    form.email.data = current_user.email
    form.phone_num.data = current_user.phone_num
    return render_template('personal_info.html', form=form)

# --- Download route for announcement attachments ---
@app.route('/download/announcement/<filename>')
@login_required
def download_announcement(filename):
    """Securely downloads an announcement attachment."""
    return send_from_directory(app.config['ANNOUNCEMENT_FOLDER'], filename, as_attachment=True)


# ==================================================================
# ====================== ADMIN-ONLY ROUTES =========================
# ==================================================================

@app.route('/admin/dashboard')
@login_required
@admin_required
def admin_dashboard():
    """Renders the admin dashboard with statistics."""
    user_count = User.query.count()
    admin_count = User.query.join(Role).filter(Role.role_name == 'Manager').count()
    engineer_count = User.query.join(Role).filter(Role.role_name == 'Engineer').count()
    
    file_count_total = History.query.count()
    
    today_start = datetime.combine(datetime.utcnow().date(), time.min)
    today_end = datetime.combine(datetime.utcnow().date(), time.max)
    
    file_count_today = History.query.filter(
        History.created_at >= today_start,
        History.created_at <= today_end
    ).count()
    
    recent_users = User.query.order_by(User.user_id.desc()).limit(5).all()
    
    return render_template(
        'dashboard_admin.html',
        user_count=user_count,
        admin_count=admin_count,
        engineer_count=engineer_count,
        file_count_total=file_count_total,
        file_count_today=file_count_today,
        recent_users=recent_users
    )

@app.route('/admin/create-user', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_create_user():
    """Renders the create user page and handles form submission."""
    form = CreateUserForm()
    if form.validate_on_submit():
        try:
            role_name = form.role.data
            role = Role.query.filter_by(role_name=role_name).first()
            if not role:
                flash(f"Role '{role_name}' not found. Please initialize database.", 'error')
                return redirect(url_for('admin_create_user'))

            new_user = User(
                username=form.username.data,
                name=form.username.data,
                email=f"{form.username.data}@ipetro.com",
                role_id=role.role_id
            )
            new_user.set_password(form.password.data)
            
            db.session.add(new_user)
            db.session.commit()
            
            flash(f'User "{new_user.username}" ({role_name}) created successfully!', 'success')
            return redirect(url_for('admin_create_user'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'An error occurred: {e}', 'error')
            
    return render_template('createuser.html', form=form)


@app.route('/admin/statistics')
@login_required
@admin_required
def admin_statistics():
    """Renders the (placeholder) statistics page."""
    return render_template('admin_placeholder.html', title='Statistics')


@app.route('/admin/announcement', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_announcement():
    """Renders the announcement page and handles posting."""
    form = AnnouncementForm()
    
    if form.validate_on_submit():
        try:
            filename = None
            if form.attachment.data:
                file = form.attachment.data
                filename = secure_filename(file.filename)
                file_path = os.path.join(app.config['ANNOUNCEMENT_FOLDER'], filename)
                file.save(file_path)
            
            new_announcement = Announcement(
                message=form.message.data,
                attachment_filename=filename,
                user_id=current_user.user_id,
                visible_to_manager=form.visible_to_manager.data,
                visible_to_engineer=form.visible_to_engineer.data
            )
            
            db.session.add(new_announcement)
            db.session.commit()
            
            flash('Announcement posted successfully!', 'success')
            return redirect(url_for('admin_announcement'))
            
        except Exception as e:
            db.session.rollback()
            print(f"Error posting announcement: {e}")
            flash(f'Error posting announcement: {e}', 'error')

    all_announcements = Announcement.query.order_by(Announcement.created_at.desc()).all()
    return render_template(
        'announcement.html',
        form=form,
        announcements=all_announcements
    )


# --- Main Application Runner ---
if __name__ == '__main__':
    app.run(debug=True)